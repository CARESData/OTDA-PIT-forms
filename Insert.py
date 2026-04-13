import csv
from openpyxl import load_workbook
import shutil
import os

# files to be used
# dir_data = r"D:\Data Source\PIT Numbers for OTDA by County\Data"
# dir_template = r"D:\Data Source\PIT Numbers for OTDA by County\OTDA Template"


def loop_through_csv(dir_data, dir_template):
    
    # loop through all the OTDA template files in the directory
    #  and fill in the data from the corresponding csv file
    for file_name in os.listdir(dir_template):
        #  current file name is in the format "NY-CountyName-OTDA Template.xlsx"
        f_temp = os.path.join(dir_template, file_name)
       
        # check the file is a file and it begins with "NY-"
        if os.path.isfile(f_temp) and file_name.startswith('NY-'):
            # get the coc from the brginning of file
            coc = file_name[:6]
            # add coc to string to get the corresponding csv file name
            file_data = os.path.join(dir_data, coc + "_filtered.csv")
            #  load the excel template file
            wb = load_workbook(f_temp)

            # open csv file and loop through each row to get the sheet name, 
            # row number, column number, and value to be entered in the cell
            with open(file_data, newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    # select  
                    sheet_name = row["County"]
                    row_num = int(float(row["row"]))
                    col_raw = int(float(row["col"]))
                    value =  row["Amount"]
                    amount = int(float(value))
                    
                    # select the worksheet
                    ws = wb[sheet_name]
                    # enter the value in the cell
                    ws.cell(row_num, col_raw, amount)


            # Save changes
            wb.save(f_temp)
    print("Step 3 done")

# main_2()
# loop_through_csv()




    